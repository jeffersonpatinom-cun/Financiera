"""Exporta a Excel la base de gestion de asesores de agosto 2026.

Acompana al Informe_Ejecutivo_Cierre_Agosto_2026.docx: el informe reporta el
equipo en agregado y sin nombres; esta base sí los trae, porque es el archivo de
trabajo del coordinador que los tiene a cargo.

CORRECCION 2026-09-03 (fe de erratas). La version anterior tomaba el universo de
Asesor_Unico, que esta disenado para NUNCA quedar vacio: cuando nadie tipifico,
cae al usuario que modifico el registro o al propietario de la cartera. Resultado:
se contaban como gestion del equipo 42.768 tipificaciones hechas por el bot
CUN DIGITAL, y como recaudo del equipo $1.849,5 MM de personas que nadie gestiono.

Reglas de calculo corregidas:
  * Gestion  : UNA FILA POR TIPIFICACION del historico
               [ZOHO].[CRM].[Historico_tipificacion_contact], excluyendo los bots
               CUN DIGITAL y PENAGOS. El asesor es Hecho_por, no Asesor_Unico.
               Antes se contaba una fila por credito cuya ULTIMA tipificacion caia
               en agosto, que no es lo mismo que una gestion.
  * Pago     : GESTION_PAGO_POST_MARCA = 1, es decir Fecha_de_pago en agosto sobre
               una persona efectivamente gestionada Y con el pago posterior a la
               primera gestion. Un pago anterior a que el asesor tocara el caso no
               es atribuible a su gestion.
  * Asignacion: la hoja Sin_Asignar SI usa Asesor_Unico. Para "de quien es la
               cartera" ese campo es el correcto; el error era usarlo para medir
               quien gestiono.
  * El dinero viene como 'CO$ 351,576.50' y sin sanear suma 0 en silencio.

Hojas:
  1. Fe_de_erratas        : cifra publicada contra cifra corregida, con la causa.
  2. Base_Gestion_Agosto  : las 22.941 gestiones humanas del mes, una por movimiento.
  3. Base_Pagos_Agosto    : los pagos de agosto atribuibles a la gestion.
  4. Resumen_Asesor       : gestiones, personas, pagos y efectividad por asesor.
  5. Resumen_Tipificacion : en que termina la gestion.
  6. Sin_Asignar          : estudiantes DE LA META vigente sin asesor responsable.
  7. Diccionario          : que significa cada columna.

Uso:  .venv/Scripts/python.exe exportar_base_gestion_agosto.py
"""
import sys

import pandas as pd
import pyodbc
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SALIDA = "Base_Gestion_Asesores_Agosto_2026.xlsx"
DESDE, HASTA = "2026-08-01", "2026-09-01"

CONN = (
    "DRIVER={ODBC Driver 18 for SQL Server};"
    "SERVER=172.16.1.33;DATABASE=CUN_REPOSITORIO;"
    "Trusted_Connection=yes;TrustServerCertificate=yes;"
)

HEX_MARINO = "0C2340"
HEX_ZEBRA = "F8F9FA"

# Saneo del dinero del CRM: 'CO$ 351,576.50' -> 351576.50
DINERO = ("TRY_CONVERT(DECIMAL(18,2), "
          "REPLACE(REPLACE(REPLACE({0},'CO$',''),',',''),' ',''))")

# Predicado de bot, identico al del SP.
NO_BOT = ("UPPER(e.Hecho_por) NOT LIKE '%CUN DIGITAL%' "
          "AND UPPER(e.Hecho_por) NOT LIKE '%PENAGOS%'")
FECHA_H = ("COALESCE(TRY_CONVERT(datetime, e.Hora_de_modificación, 103), "
           "TRY_CONVERT(datetime, e.Hora_de_creación, 103))")

# ── Gestion: una fila por tipificacion humana del mes ────────────────────────
# El JOIN a Cartera_CUN por Id trae el contexto del registro tipificado. La
# tabla materializada entra solo por el dato academico y de riesgo, que no vive
# en el CRM.
SQL_BASE = f"""
SELECT
    UPPER(LTRIM(RTRIM(CONVERT(varchar(200), e.Hecho_por))))   AS ASESOR,
    {FECHA_H}                                                 AS FECHA_GESTION,
    LTRIM(RTRIM(c.[Número_de_identificación]))                AS IDENTIFICACION,
    c.Documento_Cartera_CUN                                   AS REGISTRO,
    c.Número_de_crédito                                       AS NUMERO_CREDITO,
    c.Periodo                                                 AS PERIODO,
    c.Programa_académico                                      AS PROGRAMA,
    c.Estado_cartera                                          AS ESTADO_CARTERA,
    NULLIF(LTRIM(RTRIM(CONVERT(varchar(200), e.Tipificación_anterior))),'')
                                                              AS TIPIFICACION_ANTERIOR,
    NULLIF(LTRIM(RTRIM(CONVERT(varchar(200), e.Tipificación_nueva))),'')
                                                              AS TIPIFICACION_NUEVA,
    LTRIM(RTRIM(G.Asesor_Unico))                              AS ASESOR_ASIGNADO,
    TRY_CONVERT(date, G.Fecha_de_pago, 103)                   AS FECHA_PAGO,
    {DINERO.format('G.Valor_pagado')}                         AS VALOR_PAGADO,
    G.GESTION_PAGO_POST_MARCA                                 AS PAGO_ATRIBUIBLE,
    {DINERO.format('G.Valor_total')}                          AS VALOR_CUOTA,
    {DINERO.format('G.Deuda_total')}                          AS DEUDA_TOTAL,
    G.Días_de_mora                                            AS DIAS_MORA,
    G.CLASIFICACION_CARTERA                                   AS CLASIFICACION_CARTERA,
    G.MARCA_ACADEMICA_GESTION                                 AS MARCA_ACADEMICA,
    G.RES_PERFIL_RIESGO                                       AS PERFIL_RIESGO,
    c.Celular                                                 AS CELULAR,
    c.Correo_electrónico                                      AS CORREO
FROM [ZOHO].[CRM].[Historico_tipificacion_contact] e
JOIN ZOHO.CRM.Cartera_CUN c
      ON CONVERT(varchar(30), c.Id) = CONVERT(varchar(30), e.Cartera_CUN)
LEFT JOIN Financiera.Cartera_CUN_Asesor_Unico G
      ON CONVERT(varchar(30), G.Id) = CONVERT(varchar(30), c.Id)
WHERE e.Hecho_por IS NOT NULL AND {NO_BOT}
  AND c.[Número_de_identificación] IS NOT NULL AND c.[Número_de_identificación] <> ''
  AND {FECHA_H} >= '{DESDE}' AND {FECHA_H} < '{HASTA}'
ORDER BY ASESOR, FECHA_GESTION;
"""

# ── Pago: solo lo atribuible a la gestion ────────────────────────────────────
# GESTION_PAGO_POST_MARCA ya exige que exista gestion real Y que el pago sea
# posterior a la primera. Los pagos que quedan fuera van tipificados en la hoja
# de fe de erratas, para que se vea cuanto y por que se excluyo.
SQL_PAGOS = f"""
SELECT
    G.GESTION_ASESOR                                      AS ASESOR,
    TRY_CONVERT(date, G.Fecha_de_pago, 103)               AS FECHA_PAGO,
    LTRIM(RTRIM(G.Número_de_identificación))              AS IDENTIFICACION,
    G.Documento_Cartera_CUN                               AS REGISTRO,
    G.Número_de_crédito                                   AS NUMERO_CREDITO,
    G.Periodo                                             AS PERIODO,
    G.Programa_académico                                  AS PROGRAMA,
    G.Estado_cartera                                      AS ESTADO_CARTERA,
    NULLIF(LTRIM(RTRIM(G.Tipificación_a_marcar)),'')      AS TIPIFICACION_VIGENTE,
    {DINERO.format('G.Valor_pagado')}                     AS VALOR_PAGADO,
    {DINERO.format('G.Valor_total')}                      AS VALOR_CUOTA,
    G.GESTION_FECHA_PRIMERA                               AS PRIMERA_GESTION,
    G.GESTION_FECHA_ULTIMA                                AS ULTIMA_GESTION,
    CASE WHEN G.GESTION_FECHA_ULTIMA >= '{DESDE}'
          AND G.GESTION_FECHA_ULTIMA <  '{HASTA}'
         THEN 'SI' ELSE 'NO' END                          AS GESTIONADO_EN_AGOSTO
FROM Financiera.Cartera_CUN_Asesor_Unico G
WHERE G.GESTION_PAGO_POST_MARCA = 1
  AND TRY_CONVERT(date, G.Fecha_de_pago, 103) >= '{DESDE}'
  AND TRY_CONVERT(date, G.Fecha_de_pago, 103) <  '{HASTA}'
  AND {DINERO.format('G.Valor_pagado')} > 0
ORDER BY ASESOR, FECHA_PAGO;
"""

# Regla de negocio acordada con la Coordinacion (2026-09-02): la hoja Sin_Asignar
# lleva UNICAMENTE estudiantes que estan en la meta vigente y que en la base de
# gestion no tienen asesor responsable ('Reasignar en CRM' o 'Sin asignar').
# Sin el cruce contra la meta serian 84.179 registros de 28.619 personas, pero la
# mayoria es cartera fuera de meta que no le corresponde reasignar a Cartera.
# Aqui Asesor_Unico SI es el campo correcto: la pregunta es de asignacion.
META_VIGENTE = "202609"

SQL_SIN_ASIGNAR = f"""
WITH META AS (
    SELECT LTRIM(RTRIM(IDENTIFICACION))          AS ID,
           COUNT(*)                              AS OBLIGACIONES_EN_META,
           SUM(CAST(TOTAL AS DECIMAL(18,2)))     AS SALDO_EN_META,
           MIN([Asignacion Q])                   AS CUARTIL_MAS_URGENTE
    FROM Financiera.Cartera_Meta_Comercial_Historico
    WHERE Meta_2026 LIKE '%{META_VIGENTE}%'
    GROUP BY LTRIM(RTRIM(IDENTIFICACION))
)
SELECT
    LTRIM(RTRIM(G.Asesor_Unico))                          AS ESTADO_ASIGNACION,
    LTRIM(RTRIM(G.Número_de_identificación))              AS IDENTIFICACION,
    M.OBLIGACIONES_EN_META                                AS OBLIGACIONES_EN_META,
    CAST(M.SALDO_EN_META AS DECIMAL(18,2))                AS SALDO_EN_META,
    M.CUARTIL_MAS_URGENTE                                 AS CUARTIL_MAS_URGENTE,
    G.GESTION_MARCA                                       AS TUVO_GESTION_ALGUNA_VEZ,
    G.Documento_Cartera_CUN                               AS REGISTRO,
    G.Número_de_crédito                                   AS NUMERO_CREDITO,
    G.Periodo                                             AS PERIODO,
    G.Programa_académico                                  AS PROGRAMA,
    G.Estado_cartera                                      AS ESTADO_CARTERA,
    NULLIF(LTRIM(RTRIM(G.Tipificación_a_marcar)),'')      AS ULTIMA_TIPIFICACION,
    {DINERO.format('G.Valor_total')}                      AS VALOR_CUOTA,
    {DINERO.format('G.Deuda_total')}                      AS DEUDA_TOTAL,
    G.Días_de_mora                                        AS DIAS_MORA,
    G.CLASIFICACION_CARTERA                               AS CLASIFICACION_CARTERA,
    G.MARCA_ACADEMICA_GESTION                             AS MARCA_ACADEMICA,
    G.Celular                                             AS CELULAR,
    G.Correo_electrónico                                  AS CORREO
FROM Financiera.Cartera_CUN_Asesor_Unico G
JOIN META M ON M.ID = LTRIM(RTRIM(G.Número_de_identificación))
WHERE G.Asesor_Unico IN ('Reasignar en CRM','Sin asignar')
ORDER BY M.SALDO_EN_META DESC, IDENTIFICACION;
"""

# Lo que quedo FUERA del recaudo atribuible, y por que.
SQL_EXCLUIDO = f"""
SELECT
    CASE WHEN G.GESTION_MARCA = 0
         THEN 'Pago de persona que nadie gestiono'
         ELSE 'Pago anterior a la primera gestion' END     AS MOTIVO_EXCLUSION,
    COUNT(*)                                               AS PAGOS,
    COUNT(DISTINCT LTRIM(RTRIM(G.Número_de_identificación))) AS ESTUDIANTES,
    CAST(SUM({DINERO.format('G.Valor_pagado')}) / 1000000.0 AS DECIMAL(18,1)) AS VALOR_MM
FROM Financiera.Cartera_CUN_Asesor_Unico G
WHERE G.Asesor_Unico NOT IN ('Reasignar en CRM','Sin asignar')
  AND NULLIF(LTRIM(RTRIM(G.Asesor_Unico)),'') IS NOT NULL
  AND G.GESTION_PAGO_POST_MARCA = 0
  AND TRY_CONVERT(date, G.Fecha_de_pago, 103) >= '{DESDE}'
  AND TRY_CONVERT(date, G.Fecha_de_pago, 103) <  '{HASTA}'
  AND {DINERO.format('G.Valor_pagado')} > 0
GROUP BY CASE WHEN G.GESTION_MARCA = 0
              THEN 'Pago de persona que nadie gestiono'
              ELSE 'Pago anterior a la primera gestion' END;
"""

DICCIONARIO = [
    ("ASESOR", "Quien EJECUTO la gestion (Hecho_por del historico de tipificacion). No es Asesor_Unico: ese campo dice de quien es la cartera, no quien la trabajo."),
    ("ASESOR_ASIGNADO", "Asesor_Unico: el responsable asignado del registro. Puede diferir de ASESOR cuando gestiona alguien distinto al dueno."),
    ("FECHA_GESTION", "Momento de la tipificacion. Una fila = una gestion. Excluye las hechas por CUN DIGITAL y PENAGOS, que son cuentas de sistema."),
    ("IDENTIFICACION", "Documento del estudiante. Llave para agrupar por persona."),
    ("REGISTRO", "Identificador del registro en el CRM: identificacion-periodo-credito."),
    ("NUMERO_CREDITO", "Obligacion. Una fila es una gestion sobre una obligacion, NO una persona."),
    ("TIPIFICACION_ANTERIOR / NUEVA", "De que tipificacion a cual se movio el registro en esa gestion."),
    ("FECHA_PAGO / VALOR_PAGADO", "Pago registrado en el CRM. Ojo: es lo que el asesor marco, no la caja."),
    ("PAGO_ATRIBUIBLE", "1 cuando el pago es posterior a la PRIMERA gestion de esa persona. Un pago anterior a que el asesor tocara el caso no es fruto de su gestion."),
    ("PRIMERA_GESTION / ULTIMA_GESTION", "Solo en Base_Pagos_Agosto: rango de gestion humana sobre esa persona, en cualquier mes."),
    ("GESTIONADO_EN_AGOSTO", "Solo en Base_Pagos_Agosto: si la ultima gestion de esa persona cayo dentro del mes. Cuando dice NO, el pago viene de gestion de un mes anterior."),
    ("Gestionadas_que_pagaron", "Personas gestionadas en agosto que registraron pago atribuible en agosto, en cualquiera de sus obligaciones. Numerador de la efectividad."),
    ("VALOR_CUOTA / DEUDA_TOTAL", "Valor de la cuota y deuda total del estudiante segun el CRM."),
    ("DIAS_MORA", "Dias de mora de la obligacion."),
    ("CLASIFICACION_CARTERA", "Clasificacion de la cartera (CUENTAS POR COBRAR, CXC REFINANCIADO, etc.)."),
    ("MARCA_ACADEMICA", "Marca que enruta la gestion segun el vinculo academico del deudor."),
    ("PERFIL_RIESGO", "Perfil crediticio del deudor. Riesgo Alto o Regular es señal adversa."),
    ("ESTADO_ASIGNACION", "Solo en Sin_Asignar: si el registro esta 'Reasignar en CRM' o 'Sin asignar'."),
    ("TUVO_GESTION_ALGUNA_VEZ", "Solo en Sin_Asignar: 1 si esa persona fue gestionada alguna vez pese a no tener responsable asignado hoy."),
    ("OBLIGACIONES_EN_META / SALDO_EN_META", "Solo en Sin_Asignar: cuantas cuotas y cuanto saldo tiene esa persona en la meta vigente. Ordena la hoja de mayor a menor saldo."),
    ("CUARTIL_MAS_URGENTE", "Solo en Sin_Asignar: el cuartil mas antiguo entre las obligaciones que la persona tiene en la meta (Q1 es lo mas vencido)."),
]

# Cifras que ya recibio la Coordinacion el 2026-09-02, para la hoja de erratas.
#
# Se separan DOS causas distintas, porque no tienen la misma gravedad:
#   ERROR DE CALCULO  : la cifra estaba mal. Trabajo del bot atribuido a asesores.
#   CAMBIO DE CRITERIO: la cifra estaba bien para lo que medía, pero medía algo
#                       distinto de lo que su nombre sugería. El recaudo pasa a
#                       exigir gestion previa; antes contaba cualquier pago de
#                       cartera asignada, hubiera o no gestion detras.
PUBLICADO = {
    "Gestiones de agosto": (61767, "Error de calculo",
        "Se contaba una fila por credito cuya ultima tipificacion caia en agosto, "
        "y 42.768 de esas tipificaciones las hizo el bot CUN DIGITAL, no un asesor."),
    "Personas gestionadas": (18716, "Sin cambio relevante",
        "-1,2%. La cifra publicada ya era correcta en la practica."),
    "Asesores activos": (18, "Error de calculo",
        "Cuatro asesores figuraban activos por trabajo que ejecuto el bot."),
    "Pagos registrados": (19002, "Cambio de criterio",
        "Antes: cualquier pago de cartera asignada. Ahora: solo pagos de personas "
        "gestionadas y posteriores a la primera gestion."),
    "Estudiantes que pagaron": (11804, "Cambio de criterio",
        "Misma causa que los pagos."),
    "Valor recaudado ($MM)": (5483.9, "Cambio de criterio",
        "$1.849,5 MM eran de personas que nadie gestiono y $1.292,4 MM eran pagos "
        "anteriores a la primera gestion del asesor."),
    "Gestionadas que pagaron": (6011, "Cambio de criterio",
        "Bajo el criterio anterior (cualquier pago del mes) son 5.974, casi identico "
        "a lo publicado. La baja se debe a exigir que el pago sea posterior a la gestion."),
    "Efectividad (%)": (32.1, "Cambio de criterio",
        "Bajo el criterio anterior sigue siendo 32,3%. Con pago atribuible baja a este valor."),
}


def formatear(ws, df, congelar=True, ancho_max=38):
    relleno = PatternFill("solid", fgColor=HEX_MARINO)
    fuente = Font(name="Montserrat", bold=True, color="FFFFFF", size=10)
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j)
        c.fill = relleno
        c.font = fuente
        c.alignment = Alignment(vertical="center", horizontal="left")
        largo = df[col].head(300).astype(str).str.len().max()
        largo = 0 if pd.isna(largo) else int(largo)
        ancho = max(len(str(col)), largo) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(ancho, ancho_max)
    ws.row_dimensions[1].height = 22
    if congelar:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


def main():
    print("Conectando a 172.16.1.33 / CUN_REPOSITORIO ...")
    with pyodbc.connect(CONN) as cn:
        cn.timeout = 900
        base = pd.read_sql(SQL_BASE, cn)
        pagos = pd.read_sql(SQL_PAGOS, cn)
        sin_asignar = pd.read_sql(SQL_SIN_ASIGNAR, cn)
        excluido = pd.read_sql(SQL_EXCLUIDO, cn)
    print(f"  gestiones humanas de agosto : {len(base):,} filas")
    print(f"  pagos atribuibles           : {len(pagos):,} filas")
    print(f"  sin asignar                 : {len(sin_asignar):,} filas")

    # ---- Resumen por asesor.
    # FULL OUTER (via merge how='outer'): un asesor puede tener pagos atribuibles
    # sin haber gestionado en agosto, cuando la gestion que los origino fue de un
    # mes anterior. Con un LEFT esos pagos desaparecerian del resumen y el total
    # por asesor no cuadraria con el total del equipo.
    resumen = (base.groupby("ASESOR")
               .agg(Gestiones=("NUMERO_CREDITO", "size"),
                    Personas_gestionadas=("IDENTIFICACION", "nunique"))
               .reset_index())
    pag = (pagos.groupby("ASESOR")
           .agg(Pagos_atribuibles=("NUMERO_CREDITO", "size"),
                Personas_con_pago=("IDENTIFICACION", "nunique"),
                Valor_pagado=("VALOR_PAGADO", "sum"))
           .reset_index())

    ids_pagaron = set(pagos["IDENTIFICACION"])
    cruce = (base[base["IDENTIFICACION"].isin(ids_pagaron)]
             .groupby("ASESOR")["IDENTIFICACION"].nunique()
             .rename("Gestionadas_que_pagaron").reset_index())

    resumen = (resumen.merge(pag, on="ASESOR", how="outer")
                      .merge(cruce, on="ASESOR", how="left")
                      .fillna({"Gestiones": 0, "Personas_gestionadas": 0,
                               "Pagos_atribuibles": 0, "Personas_con_pago": 0,
                               "Valor_pagado": 0, "Gestionadas_que_pagaron": 0}))
    for c in ("Gestiones", "Personas_gestionadas", "Pagos_atribuibles",
              "Personas_con_pago", "Gestionadas_que_pagaron"):
        resumen[c] = resumen[c].astype(int)
    # replace(0, nan) y no pd.NA: el asesor sin gestion en agosto (pero con pagos
    # de gestion previa) dejaria una division por cero, y pd.NA rompe el round.
    resumen["Efectividad_pct"] = (100 * resumen["Gestionadas_que_pagaron"]
                                  / resumen["Personas_gestionadas"]
                                  .replace(0, float("nan"))).round(2)
    resumen["Valor_pagado_MM"] = (resumen["Valor_pagado"] / 1e6).round(1)
    resumen["Pct_de_la_gestion"] = (100 * resumen["Gestiones"]
                                    / resumen["Gestiones"].sum()).round(2)
    resumen = resumen.drop(columns=["Valor_pagado"]).sort_values(
        "Valor_pagado_MM", ascending=False)

    gest_pago = base[base["IDENTIFICACION"].isin(ids_pagaron)]["IDENTIFICACION"].nunique()
    total = pd.DataFrame([{
        "ASESOR": "TOTAL EQUIPO",
        "Gestiones": len(base),
        "Personas_gestionadas": base["IDENTIFICACION"].nunique(),
        "Pagos_atribuibles": len(pagos),
        "Personas_con_pago": pagos["IDENTIFICACION"].nunique(),
        "Gestionadas_que_pagaron": gest_pago,
        "Efectividad_pct": round(100 * gest_pago / base["IDENTIFICACION"].nunique(), 2),
        "Valor_pagado_MM": round(pagos["VALOR_PAGADO"].sum() / 1e6, 1),
        "Pct_de_la_gestion": 100.0}])
    resumen = pd.concat([resumen, total], ignore_index=True)

    # ---- Resumen por tipificacion (la NUEVA, que es la que aplico el asesor)
    tipif = (base.assign(TIPIFICACION_NUEVA=base["TIPIFICACION_NUEVA"]
                         .fillna("(sin tipificar)"))
             .groupby("TIPIFICACION_NUEVA")
             .agg(Gestiones=("NUMERO_CREDITO", "size"),
                  Personas=("IDENTIFICACION", "nunique"))
             .reset_index()
             .sort_values("Gestiones", ascending=False))
    tipif["Pct"] = (100 * tipif["Gestiones"] / tipif["Gestiones"].sum()).round(2)

    # ---- Hoja de fe de erratas
    corregido = {
        "Gestiones de agosto": len(base),
        "Personas gestionadas": base["IDENTIFICACION"].nunique(),
        "Asesores activos": base["ASESOR"].nunique(),
        "Gestionadas que pagaron": gest_pago,
        "Pagos registrados": len(pagos),
        "Estudiantes que pagaron": pagos["IDENTIFICACION"].nunique(),
        "Valor recaudado ($MM)": round(pagos["VALOR_PAGADO"].sum() / 1e6, 1),
        "Efectividad (%)": round(100 * gest_pago / base["IDENTIFICACION"].nunique(), 1),
    }
    erratas = pd.DataFrame(
        [{"Cifra": k, "Publicado_02sep": v[0], "Corregido": corregido[k],
          "Diferencia": round(corregido[k] - v[0], 1), "Tipo": v[1], "Causa": v[2]}
         for k, v in PUBLICADO.items()])

    diccionario = pd.DataFrame(DICCIONARIO, columns=["Columna", "Significado"])

    print(f"Escribiendo {SALIDA} ...")
    with pd.ExcelWriter(SALIDA, engine="openpyxl") as xl:
        for nombre, df, congelar in [
            ("Fe_de_erratas", erratas, False),
            ("Base_Gestion_Agosto", base, True),
            ("Base_Pagos_Agosto", pagos, True),
            ("Resumen_Asesor", resumen, False),
            ("Resumen_Tipificacion", tipif, False),
            ("Recaudo_Excluido", excluido, False),
            ("Sin_Asignar", sin_asignar, True),
            ("Diccionario", diccionario, False),
        ]:
            df.to_excel(xl, sheet_name=nombre, index=False)
            formatear(xl.sheets[nombre], df, congelar=congelar)

        we = xl.sheets["Fe_de_erratas"]
        we.column_dimensions["F"].width = 96
        for i in range(2, len(erratas) + 2):
            we.cell(row=i, column=6).alignment = Alignment(wrap_text=True, vertical="top")

        # zebra suave + total en negrita en el resumen por asesor
        wr = xl.sheets["Resumen_Asesor"]
        for i in range(2, len(resumen) + 2):
            if i % 2 == 1:
                for j in range(1, len(resumen.columns) + 1):
                    wr.cell(row=i, column=j).fill = PatternFill("solid", fgColor=HEX_ZEBRA)
        for j in range(1, len(resumen.columns) + 1):
            wr.cell(row=len(resumen) + 1, column=j).font = Font(bold=True)

        wd = xl.sheets["Diccionario"]
        wd.column_dimensions["A"].width = 34
        wd.column_dimensions["B"].width = 104
        for i in range(2, len(diccionario) + 2):
            wd.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    print(f"OK -> {SALIDA}")
    print(f"   Gestiones humanas    : {len(base):,}")
    print(f"   Asesores con gestion : {base['ASESOR'].nunique()}")
    print(f"   Personas gestionadas : {base['IDENTIFICACION'].nunique():,}")
    print(f"   Pagos atribuibles    : {len(pagos):,}  "
          f"(${pagos['VALOR_PAGADO'].sum()/1e6:,.1f} MM)")
    print(f"   Efectividad equipo   : {resumen.iloc[-1]['Efectividad_pct']}%")
    print(f"   Sin asignar (meta)   : {len(sin_asignar):,} registros, "
          f"{sin_asignar['IDENTIFICACION'].nunique():,} personas, "
          f"${sin_asignar.drop_duplicates('IDENTIFICACION')['SALDO_EN_META'].sum()/1e6:,.1f} MM en la meta")


if __name__ == "__main__":
    main()
