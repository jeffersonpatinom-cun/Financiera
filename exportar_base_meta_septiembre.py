"""Exporta la base de la Meta Comercial de septiembre 2026 a Excel.

Universo: [Financiera].[Cartera_Meta_Comercial_Historico] con marca '202609' en
Meta_2026 (56.269 obligaciones). Es la misma base del informe ejecutivo.

Hojas:
  1. Base_Meta_Septiembre : las 51 columnas de la tabla + 4 campos derivados de apoyo.
  2. Resumen              : los cortes del informe, para cuadrar contra el documento.
  3. Diccionario          : que significa cada columna.

Uso:  .venv/Scripts/python.exe exportar_base_meta_septiembre.py
"""
import sys

import pandas as pd
import pyodbc
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SALIDA = "Base_Meta_Comercial_Septiembre_2026.xlsx"
MARCA = "202609"

CONN = (
    "DRIVER={ODBC Driver 18 for SQL Server};"
    "SERVER=172.16.1.33;DATABASE=CUN_REPOSITORIO;"
    "Trusted_Connection=yes;TrustServerCertificate=yes;"
)

# ---- Paleta institucional CUN
HEX_MARINO = "0C2340"
HEX_MANZANA = "84BD00"
HEX_ZEBRA = "F8F9FA"
HEX_GRIS = "898D8D"

SQL_BASE = f"""
SELECT
    H.*,
    -- ---- campos derivados de apoyo a la gestion (no existen en la tabla) ----
    CASE WHEN H.Anio_Mes_Ingreso = '{MARCA}' THEN 'NUEVO EN SEPTIEMBRE'
         ELSE 'ARRASTRE' END                                        AS ORIGEN_META,
    LEN(H.Meta_2026) - LEN(REPLACE(H.Meta_2026, ',', '')) + 1       AS MESES_EN_META,
    CASE WHEN CAST(H.GR360MAS  AS DECIMAL(18,2)) > 0 THEN '7. +360 dias'
         WHEN CAST(H.GR151A360 AS DECIMAL(18,2)) > 0 THEN '6. 151-360'
         WHEN CAST(H.GR121A150 AS DECIMAL(18,2)) > 0 THEN '5. 121-150'
         WHEN CAST(H.GR91A120  AS DECIMAL(18,2)) > 0 THEN '4. 91-120'
         WHEN CAST(H.GR61A90   AS DECIMAL(18,2)) > 0 THEN '3. 61-90'
         WHEN CAST(H.GR31A60   AS DECIMAL(18,2)) > 0 THEN '2. 31-60'
         WHEN CAST(H.GR1A30    AS DECIMAL(18,2)) > 0 THEN '1. 1-30'
         ELSE '0. sin edad de mora' END                             AS EDAD_MORA,
    TRY_CONVERT(date, H.FECHA_VENCIMIENTO, 103)                     AS FECHA_VENCIMIENTO_DATE
FROM Financiera.Cartera_Meta_Comercial_Historico H
WHERE H.Meta_2026 LIKE '%{MARCA}%'
ORDER BY H.[Asignacion Q], H.IDENTIFICACION, H.NUMERO_CREDITO;
"""

DICCIONARIO = [
    ("FECHA_CORTE", "Fecha de corte de la cartera con la que se refresco la fila."),
    ("PERIODO", "Periodo academico de la obligacion (cod_periodo)."),
    ("IDENTIFICACION", "Documento del estudiante deudor. Llave para agrupar por persona."),
    ("NOMBRE / EMAIL / TEL_CELULAR / WHATSAPP", "Datos de contacto vigentes, refrescados cada mes."),
    ("NUMERO_CREDITO", "Llave unica de la obligacion. Una fila = una cuota, NO una persona."),
    ("FECHA_VENCIMIENTO", "Vencimiento de la cuota, en texto dd/MM/yyyy."),
    ("FECHA_VENCIMIENTO_DATE", "Derivado: el mismo vencimiento como fecha real, para ordenar y filtrar."),
    ("VALOR_ORIGINAL", "Valor inicial de la cuota."),
    ("GR1A30 ... GR360MAS", "Saldo distribuido por rango de dias de mora."),
    ("TOTAL", "Saldo vigente de la obligacion. Es la columna que suma la meta."),
    ("Asignacion Q", "Cuartil de gestion: Q1 lo mas vencido, Q4 lo mas reciente. 25% del saldo cada uno. Se recalcula cada mes."),
    ("Ultimo_acceso_moodle", "Ultimo ingreso a plataforma academica. Vacio = sin registro de acceso."),
    ("Promedio_notas", "Promedio acumulado del estudiante (PRO_ACUMULADO). Vacio en el 14,3% de los casos."),
    ("ESTADO_ALUMNO", "Vinculo con la institucion: 1-Activo, 2-Egresado, 3-Graduado."),
    ("MARCA_ACADEMICA", "Marca que enruta la gestion de cobro. Escalera excluyente, nunca vacia."),
    ("MARCA_ACADEMICA_DETALLE", "Subcategoria de la marca; abre cada una segun el riesgo crediticio."),
    ("Meta_2026", "Bitacora de los meses en que el credito estuvo en la meta ('202606, 202607, ...')."),
    ("Anio_Mes_Ingreso", "Mes en que el credito entro por primera vez a la meta."),
    ("ORIGEN_META", "Derivado: NUEVO EN SEPTIEMBRE o ARRASTRE, segun Anio_Mes_Ingreso."),
    ("MESES_EN_META", "Derivado: cuantos meses lleva el credito en la meta. 4 = esta desde junio."),
    ("EDAD_MORA", "Derivado: rango de mora mas antiguo con saldo, a partir de las columnas GR*."),
    ("AUD_FECHA_FOTO", "Instante del primer ingreso a la meta."),
    ("AUD_FECHA_ACTUALIZACION", "Instante del ultimo refresco mensual."),
]


def autoajustar(ws, df, ancho_max=42):
    for j, col in enumerate(df.columns, start=1):
        largo = df[col].head(300).astype(str).str.len().max()
        largo = 0 if pd.isna(largo) else int(largo)
        ancho = max(len(str(col)), largo) + 2
        ws.column_dimensions[get_column_letter(j)].width = min(ancho, ancho_max)


def formatear_encabezado(ws, ncols, congelar="A2"):
    relleno = PatternFill("solid", fgColor=HEX_MARINO)
    fuente = Font(name="Montserrat", bold=True, color="FFFFFF", size=10)
    for j in range(1, ncols + 1):
        c = ws.cell(row=1, column=j)
        c.fill = relleno
        c.font = fuente
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=False)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = congelar


def main():
    print("Conectando a 172.16.1.33 / CUN_REPOSITORIO ...")
    with pyodbc.connect(CONN) as cn:
        cn.timeout = 600
        base = pd.read_sql(SQL_BASE, cn)
    print(f"  base: {len(base):,} filas x {len(base.columns)} columnas")

    # Resumen: los mismos cortes del informe ejecutivo, calculados sobre la base.
    saldo = base["TOTAL"].sum()
    resumen = pd.DataFrame(
        [["Obligaciones", len(base)],
         ["Estudiantes distintos", base["IDENTIFICACION"].nunique()],
         ["Saldo total ($)", round(saldo, 2)],
         ["Saldo total ($ millones)", round(saldo / 1e6, 1)],
         ["Ticket promedio ($)", round(base["TOTAL"].mean(), 0)],
         ["Obligaciones por estudiante", round(len(base) / base["IDENTIFICACION"].nunique(), 2)]],
        columns=["Indicador", "Valor"])

    def corte(col, etiqueta):
        g = (base.groupby(col, dropna=False)
                 .agg(Obligaciones=("NUMERO_CREDITO", "size"),
                      Estudiantes=("IDENTIFICACION", "nunique"),
                      Saldo_MM=("TOTAL", lambda x: round(x.sum() / 1e6, 1)))
                 .reset_index()
                 .rename(columns={col: etiqueta})
                 .sort_values("Obligaciones", ascending=False))
        g["Pct_saldo"] = (100 * g["Saldo_MM"] / g["Saldo_MM"].sum()).round(2)
        return g

    cortes = [("Resumen general", resumen),
              ("Por origen", corte("ORIGEN_META", "Origen")),
              ("Por cuartil de gestion", corte("Asignacion Q", "Cuartil").sort_values("Cuartil")),
              ("Por marca academica", corte("MARCA_ACADEMICA", "Marca academica")),
              ("Por detalle de marca", corte("MARCA_ACADEMICA_DETALLE", "Detalle")),
              ("Por edad de mora", corte("EDAD_MORA", "Edad de mora").sort_values("Edad de mora")),
              ("Por meses en la meta", corte("MESES_EN_META", "Meses en la meta").sort_values("Meses en la meta")),
              ("Por periodo academico", corte("PERIODO", "Periodo"))]

    diccionario = pd.DataFrame(DICCIONARIO, columns=["Columna", "Significado"])

    print(f"Escribiendo {SALIDA} ...")
    with pd.ExcelWriter(SALIDA, engine="openpyxl") as xl:
        base.to_excel(xl, sheet_name="Base_Meta_Septiembre", index=False)
        ws = xl.sheets["Base_Meta_Septiembre"]
        formatear_encabezado(ws, len(base.columns))
        autoajustar(ws, base)
        ws.auto_filter.ref = ws.dimensions

        # ---- Hoja Resumen: los bloques uno debajo del otro
        fila = 0
        for titulo, df in cortes:
            enc = pd.DataFrame([[titulo]])
            enc.to_excel(xl, sheet_name="Resumen", index=False, header=False, startrow=fila)
            df.to_excel(xl, sheet_name="Resumen", index=False, startrow=fila + 1)
            fila += len(df) + 4
        wr = xl.sheets["Resumen"]
        fila = 0
        for titulo, df in cortes:
            t = wr.cell(row=fila + 1, column=1)
            t.font = Font(name="Montserrat", bold=True, size=11, color=HEX_MARINO)
            for j in range(1, len(df.columns) + 1):
                c = wr.cell(row=fila + 2, column=j)
                c.fill = PatternFill("solid", fgColor=HEX_MARINO)
                c.font = Font(name="Montserrat", bold=True, color="FFFFFF", size=9.5)
            fila += len(df) + 4
        for j, w in enumerate([34, 16, 14, 14, 12], start=1):
            wr.column_dimensions[get_column_letter(j)].width = w

        diccionario.to_excel(xl, sheet_name="Diccionario", index=False)
        wd = xl.sheets["Diccionario"]
        formatear_encabezado(wd, 2)
        wd.column_dimensions["A"].width = 42
        wd.column_dimensions["B"].width = 96
        for i in range(2, len(diccionario) + 2):
            wd.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    print(f"OK -> {SALIDA}")
    print(f"   Base_Meta_Septiembre : {len(base):,} filas")
    print(f"   Saldo                : ${saldo:,.0f}  ({saldo/1e6:,.1f} MM)")


if __name__ == "__main__":
    main()
