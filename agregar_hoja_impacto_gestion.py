# -*- coding: utf-8 -*-
"""
Agrega la hoja "Impacto gestion" al libro Evidencia_NDB_Fantasma_<fecha>.xlsx.

Autor: Analitica financiera - Universidad CUN

Mide cuanto trabajo del equipo de cobranza se esta gastando sobre obligaciones ya
pagadas, cruzando la cartera fantasma contra Financiera.Cartera_CUN_Asesor_Unico
(la herramienta de gestion).

Igual que ajustar_evidencia_dos_poblaciones.py: ABRE el libro existente y solo agrega.
No regenera nada, para no perder las ediciones manuales.

Llave de cruce (verificada 1:1 en ambos lados):
    Cartera_Gestion  : IDENTIFICACION + '-' + PERIODO + '-' + NUMERO_CREDITO
    Asesor_Unico     : Documento_Cartera_CUN

OJO: no cruzar por Número_de_identificación a secas -- ambas tablas tienen grano
obligacion y eso produce un cartesiano N×N.

CORREGIDO 2026-09-03. La primera version de esta hoja conto como "trabajo de asesor
sobre deuda inexistente" las tipificaciones del bot CUN DIGITAL, porque el flag
GESTIONADA se basaba en Tipificación_nueva (que sale de la ultima tipificacion
cualquiera, bots incluidos) y atribuia el caso al Asesor_Unico, campo que nunca queda
vacio. Ahora la gestion se lee del historico, al grano de OBLIGACION y solo humana:
de 3.454 obligaciones que la version anterior daba por gestionadas, 1.566 las habia
tocado unicamente el robot.

Si el libro ya tiene la hoja, se REEMPLAZA (la anterior tiene cifras infladas).
"""

import sys

import pyodbc
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ARCHIVO = 'Evidencia_NDB_Fantasma_2026-08-28.xlsx'
HOJA = 'Impacto gestion'

CONN_STR = (
    'DRIVER={ODBC Driver 18 for SQL Server};'
    'SERVER=172.16.1.33;DATABASE=CUN_REPOSITORIO;'
    'Trusted_Connection=yes;TrustServerCertificate=yes;Connect Timeout=30;'
)

AZUL_MARINO = '0C2340'
AZUL_TURQ = '00859B'
GRIS_CLARO = 'F8F9FA'
GRIS_INST = '898D8D'

F_TITULO = Font(name='Montserrat', size=14, bold=True, color=AZUL_MARINO)
F_SUB = Font(name='Montserrat', size=10, italic=True, color=GRIS_INST)
F_BLOQUE = Font(name='Montserrat', size=11, bold=True, color=AZUL_MARINO)
F_HEAD = Font(name='Montserrat', size=10, bold=True, color='FFFFFF')
F_BODY = Font(name='Open Sans', size=9.5, color='222222')
F_KPI = Font(name='Montserrat', size=11, bold=True, color=AZUL_TURQ)
F_NOTA = Font(name='Open Sans', size=8.5, italic=True, color=GRIS_INST)

FILL_HEAD = PatternFill('solid', fgColor=AZUL_MARINO)
FILL_ZEBRA = PatternFill('solid', fgColor=GRIS_CLARO)
BORDE = Border(bottom=Side(style='thin', color='E0E0E0'))

# Universo fantasma: misma definicion que el resto del informe.
FANT_SQL = """
IF OBJECT_ID('tempdb..#FANT') IS NOT NULL DROP TABLE #FANT;
SELECT LTRIM(RTRIM(IDENTIFICACION)) + '-' + LTRIM(RTRIM(PERIODO)) + '-'
         + LTRIM(RTRIM(CAST(NUMERO_CREDITO AS varchar(20)))) AS DOC_KEY,
       IDENTIFICACION,
       CAST(VALOR_ORIGINAL AS DECIMAL(18,2)) AS VALOR_ORIGINAL
INTO #FANT
FROM Financiera.Cartera_Gestion
WHERE DOCUMENTO = 'NDB'
  AND CAST(VALOR_ORIGINAL AS DECIMAL(18,4)) >= 50000
  AND CAST(TOTAL AS DECIMAL(18,4)) < 1000
  AND CAST(TOTAL AS DECIMAL(18,4)) >= 0;

/* Gestion HUMANA al grano de obligacion. No se usa GESTION_MARCA de la tabla
   materializada: esa columna esta al grano de CEDULA y diria que la obligacion se
   gestiono cuando en realidad se gestiono otra cuota de la misma persona. */
IF OBJECT_ID('tempdb..#GEST_OBL') IS NOT NULL DROP TABLE #GEST_OBL;
SELECT cartera_id, TIPIF_HUMANA, GESTOR_HUMANO, TOQUES_HUMANOS
INTO #GEST_OBL
FROM (
    SELECT CONVERT(varchar(30), e.Cartera_CUN) AS cartera_id,
           CONVERT(varchar(200), e.Tipificación_nueva) AS TIPIF_HUMANA,
           UPPER(LTRIM(RTRIM(CONVERT(varchar(200), e.Hecho_por)))) AS GESTOR_HUMANO,
           COUNT(*) OVER (PARTITION BY CONVERT(varchar(30), e.Cartera_CUN)) AS TOQUES_HUMANOS,
           ROW_NUMBER() OVER (PARTITION BY CONVERT(varchar(30), e.Cartera_CUN)
                ORDER BY COALESCE(TRY_CONVERT(datetime, e.Hora_de_modificación, 103),
                                  TRY_CONVERT(datetime, e.Hora_de_creación, 103)) DESC,
                         CONVERT(varchar(30), e.Id) DESC) AS rn
    FROM [ZOHO].[CRM].[Historico_tipificacion_contact] e
    WHERE e.Cartera_CUN IS NOT NULL AND e.Hecho_por IS NOT NULL
      AND UPPER(e.Hecho_por) NOT LIKE '%CUN DIGITAL%'
      AND UPPER(e.Hecho_por) NOT LIKE '%PENAGOS%'
) x WHERE x.rn = 1;
CREATE UNIQUE CLUSTERED INDEX IX_tmp_gestobl ON #GEST_OBL (cartera_id);

IF OBJECT_ID('tempdb..#BOT_OBL') IS NOT NULL DROP TABLE #BOT_OBL;
SELECT CONVERT(varchar(30), e.Cartera_CUN) AS cartera_id, COUNT(*) AS TOQUES_BOT
INTO #BOT_OBL
FROM [ZOHO].[CRM].[Historico_tipificacion_contact] e
WHERE e.Cartera_CUN IS NOT NULL AND e.Hecho_por IS NOT NULL
  AND (UPPER(e.Hecho_por) LIKE '%CUN DIGITAL%' OR UPPER(e.Hecho_por) LIKE '%PENAGOS%')
GROUP BY CONVERT(varchar(30), e.Cartera_CUN);
CREATE UNIQUE CLUSTERED INDEX IX_tmp_botobl ON #BOT_OBL (cartera_id);

IF OBJECT_ID('tempdb..#CRUCE') IS NOT NULL DROP TABLE #CRUCE;
SELECT F.DOC_KEY, F.IDENTIFICACION, F.VALOR_ORIGINAL,
       A.Asesor_Unico AS ASESOR_ASIGNADO, G.GESTOR_HUMANO AS GESTIONADO_POR,
       A.Estado_cartera AS ESTADO_CARTERA,
       G.TIPIF_HUMANA AS TIPIFICACION, A.Fechahora_llamada AS LLAMADA,
       ISNULL(G.TOQUES_HUMANOS, 0) AS TOQUES_HUMANOS,
       ISNULL(B.TOQUES_BOT, 0)     AS TOQUES_BOT,
       CASE WHEN A.Documento_Cartera_CUN IS NULL THEN 'No esta en la herramienta'
            WHEN LTRIM(RTRIM(ISNULL(A.Asesor_Unico,''))) IN ('','Reasignar en CRM','Sin asignar','CUN DIGITAL')
                 THEN 'En la herramienta, sin asesor asignado'
            ELSE 'Asignada a un asesor' END AS SITUACION_ASIGNACION,
       /* Trabajo HUMANO sobre esta obligacion. La llamada cuenta aunque no haya
          tipificacion; el toque del bot ya no. */
       CASE WHEN G.cartera_id IS NOT NULL
              OR NULLIF(LTRIM(RTRIM(ISNULL(A.Fechahora_llamada,''))),'') IS NOT NULL
            THEN 1 ELSE 0 END AS GESTIONADA
INTO #CRUCE
FROM #FANT F
LEFT JOIN Financiera.Cartera_CUN_Asesor_Unico A
       ON LTRIM(RTRIM(A.Documento_Cartera_CUN)) = F.DOC_KEY
LEFT JOIN #GEST_OBL G ON G.cartera_id = CONVERT(varchar(30), A.Id)
LEFT JOIN #BOT_OBL  B ON B.cartera_id = CONVERT(varchar(30), A.Id);
"""

BLOQUES = [
    ('1. Donde estan las obligaciones ya pagadas',
     """SELECT SITUACION_ASIGNACION, COUNT(*) AS OBLIGACIONES,
               COUNT(DISTINCT IDENTIFICACION) AS PERSONAS,
               CAST(100.0*COUNT(*)/SUM(COUNT(*)) OVER () AS DECIMAL(5,2)) AS PCT
        FROM #CRUCE GROUP BY SITUACION_ASIGNACION ORDER BY OBLIGACIONES DESC;"""),

    ('2. Trabajo ejercido sobre deuda inexistente',
     """SELECT 'Fantasmas cargadas en la herramienta de gestion' AS INDICADOR,
               COUNT(*) AS OBLIGACIONES, COUNT(DISTINCT IDENTIFICACION) AS PERSONAS
        FROM #CRUCE WHERE SITUACION_ASIGNACION <> 'No esta en la herramienta'
        UNION ALL SELECT 'Asignadas a un asesor', COUNT(*), COUNT(DISTINCT IDENTIFICACION)
        FROM #CRUCE WHERE SITUACION_ASIGNACION = 'Asignada a un asesor'
        UNION ALL SELECT 'CON gestion HUMANA (tipificacion o llamada)', COUNT(*), COUNT(DISTINCT IDENTIFICACION)
        FROM #CRUCE WHERE GESTIONADA = 1
        UNION ALL SELECT 'Con llamada registrada', COUNT(*), COUNT(DISTINCT IDENTIFICACION)
        FROM #CRUCE WHERE NULLIF(LTRIM(RTRIM(ISNULL(LLAMADA,''))),'') IS NOT NULL
        UNION ALL SELECT 'Tocadas SOLO por el bot (no es trabajo de asesor)', COUNT(*), COUNT(DISTINCT IDENTIFICACION)
        FROM #CRUCE WHERE GESTIONADA = 0 AND TOQUES_BOT > 0;"""),

    ('3. Carga por asesor: lo asignado contra lo efectivamente trabajado',
     """WITH COLA AS (SELECT LTRIM(RTRIM(Asesor_Unico)) AS ASESOR, COUNT(*) AS CASOS_TOTALES
                      FROM Financiera.Cartera_CUN_Asesor_Unico GROUP BY LTRIM(RTRIM(Asesor_Unico))),
             FA AS (SELECT LTRIM(RTRIM(ASESOR_ASIGNADO)) AS ASESOR, COUNT(*) AS CASOS_FANTASMA,
                           COUNT(DISTINCT IDENTIFICACION) AS PERSONAS
                    FROM #CRUCE WHERE SITUACION_ASIGNACION = 'Asignada a un asesor'
                    GROUP BY LTRIM(RTRIM(ASESOR_ASIGNADO))),
             FW AS (SELECT GESTIONADO_POR AS ASESOR, COUNT(*) AS FANTASMAS_TRABAJADAS,
                           SUM(TOQUES_HUMANOS) AS TOQUES_DESPERDICIADOS
                    FROM #CRUCE WHERE GESTIONADO_POR IS NOT NULL
                    GROUP BY GESTIONADO_POR)
        SELECT ISNULL(FA.ASESOR, FW.ASESOR) AS ASESOR, C.CASOS_TOTALES AS COLA_TOTAL,
               ISNULL(FA.CASOS_FANTASMA,0) AS CASOS_FANTASMA, ISNULL(FA.PERSONAS,0) AS PERSONAS,
               ISNULL(FW.FANTASMAS_TRABAJADAS,0) AS FANTASMAS_TRABAJADAS,
               ISNULL(FW.TOQUES_DESPERDICIADOS,0) AS TOQUES_DESPERDICIADOS,
               CAST(100.0*ISNULL(FA.CASOS_FANTASMA,0)/NULLIF(C.CASOS_TOTALES,0) AS DECIMAL(5,2)) AS PCT_DE_SU_COLA
        FROM FA FULL JOIN FW ON FW.ASESOR = FA.ASESOR
        LEFT JOIN COLA C ON C.ASESOR = ISNULL(FA.ASESOR, FW.ASESOR)
        ORDER BY FANTASMAS_TRABAJADAS DESC, CASOS_FANTASMA DESC;"""),

    ('4. Que tipificaron los asesores sobre estas obligaciones',
     """SELECT LTRIM(RTRIM(TIPIFICACION)) AS TIPIFICACION, COUNT(*) AS OBLIGACIONES,
               COUNT(DISTINCT IDENTIFICACION) AS PERSONAS
        FROM #CRUCE WHERE NULLIF(LTRIM(RTRIM(ISNULL(TIPIFICACION,''))),'') IS NOT NULL
        GROUP BY LTRIM(RTRIM(TIPIFICACION)) ORDER BY OBLIGACIONES DESC;"""),

    ('5. Contraste: cartera real vs. cartera fantasma',
     """SELECT 'Cartera real (deuda vigente)' AS UNIVERSO, COUNT(*) AS OBLIGACIONES,
               SUM(CASE WHEN G.cartera_id IS NOT NULL
                          OR NULLIF(LTRIM(RTRIM(ISNULL(A.Fechahora_llamada,''))),'') IS NOT NULL
                        THEN 1 ELSE 0 END) AS CON_GESTION
        FROM Financiera.Cartera_CUN_Asesor_Unico A
        LEFT JOIN #GEST_OBL G ON G.cartera_id = CONVERT(varchar(30), A.Id)
        WHERE NOT EXISTS (SELECT 1 FROM #FANT F WHERE F.DOC_KEY = LTRIM(RTRIM(A.Documento_Cartera_CUN)))
        UNION ALL
        SELECT 'Cartera fantasma (ya pagada)', COUNT(*), SUM(GESTIONADA)
        FROM #CRUCE WHERE SITUACION_ASIGNACION <> 'No esta en la herramienta';"""),
]

NOTAS = [
    'Cruce por Documento_Cartera_CUN (identificacion-periodo-credito), llave verificada 1:1 en ambas tablas.',
    '"Con gestion HUMANA" = la obligacion tiene una tipificacion de una persona o una llamada asociada.',
    'Las tipificaciones del bot CUN DIGITAL NO cuentan como trabajo de asesor: son el 54,3% del historico y',
    'la version anterior de esta hoja las contaba, inflando el desperdicio atribuido al equipo.',
    'ASESOR_ASIGNADO responde "de quien es el caso" (Asesor_Unico); GESTIONADO_POR responde "quien lo trabajo".',
    'Los rotulos "Reasignar en CRM", "Sin asignar" y "CUN DIGITAL" no son personas: se cuentan como sin asesor.',
    'Este conteo mide OBLIGACIONES gestionadas, no tiempo. Convertirlo a horas o costo requiere un dato de',
    'productividad (duracion promedio por gestion) que debe aportar la Coordinacion de Cartera.',
]


def main():
    wb = load_workbook(ARCHIVO)
    # Antes se abortaba si la hoja existia. Ahora se reemplaza: la version del
    # 2026-08-28 quedo con cifras infladas por contar el trabajo del bot.
    if HOJA in wb.sheetnames:
        print(f'Reemplazando la hoja "{HOJA}" (la anterior contaba gestion del bot).')
        del wb[HOJA]

    cn = pyodbc.connect(CONN_STR, timeout=30)
    cn.timeout = 600
    cur = cn.cursor()
    cur.execute(FANT_SQL)
    while cur.nextset():
        pass

    ws = wb.create_sheet(HOJA, 4)          # despues de "Clientes 100 pct fantasma"
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'Impacto sobre el equipo de gestion de cobranza'
    ws['A1'].font = F_TITULO
    ws['A2'] = ('Cuanto trabajo se esta ejerciendo sobre obligaciones que ya fueron pagadas.  '
                'Fuente: Financiera.Cartera_CUN_Asesor_Unico')
    ws['A2'].font = F_SUB

    fila = 4
    anchos = {}
    for titulo, sql in BLOQUES:
        ws.cell(row=fila, column=1, value=titulo).font = F_BLOQUE
        fila += 1

        cur.execute(sql)
        cols = [c[0] for c in cur.description]
        filas = cur.fetchall()

        for j, c in enumerate(cols, start=1):
            cel = ws.cell(row=fila, column=j, value=c)
            cel.font, cel.fill = F_HEAD, FILL_HEAD
            cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            anchos[j] = max(anchos.get(j, 0), len(str(c)))
        fila += 1

        for i, row in enumerate(filas):
            for j, v in enumerate(row, start=1):
                cel = ws.cell(row=fila + i, column=j, value=v)
                cel.font, cel.border = F_BODY, BORDE
                if i % 2 == 1:
                    cel.fill = FILL_ZEBRA
                anchos[j] = max(anchos.get(j, 0), len(str(v)) if v is not None else 0)
            # La columna 2 de los bloques de indicadores es la cifra protagonista.
            if titulo.startswith('2.'):
                ws.cell(row=fila + i, column=2).font = F_KPI
        fila += len(filas) + 2
        print(f'  {titulo}: {len(filas)} filas')

    for j, w in anchos.items():
        ws.column_dimensions[ws.cell(row=5, column=j).column_letter].width = min(w + 3, 52)

    fila += 1
    for n in NOTAS:
        ws.cell(row=fila, column=1, value=n).font = F_NOTA
        fila += 1

    wb.save(ARCHIVO)
    cn.close()
    print(f'\nOK -> hoja "{HOJA}" agregada a {ARCHIVO}')


if __name__ == '__main__':
    main()
